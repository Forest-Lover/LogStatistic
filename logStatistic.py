# -*- coding: utf-8 -*-
import argparse
import json
import glob
import platform
import re
import psutil
import os
import threading
import time
import numpy as np
import pandas as pd
import subprocess
from decimal import Decimal, ROUND_HALF_UP
from collections import Counter, defaultdict
from datetime import datetime, timedelta, timezone
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.cluster import DBSCAN, KMeans
from sklearn.metrics.pairwise import cosine_similarity
from concurrent.futures import ThreadPoolExecutor, ProcessPoolExecutor, as_completed
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def memory_watcher(max_memory, check_interval):
    """
    Monitor the memory usage of the current process in a separate thread.
    :param max_memory: Maximum memory in MB.
    :param check_interval: Time interval in seconds between memory checks.
    """
    process = psutil.Process()
    start_time = time.time()
    while True:
        current_time = time.time()
        elapsed_time = current_time - start_time
        memory_usage = process.memory_info().rss / (1024 * 1024)
        if memory_usage > max_memory:
            print(f"Memory limit exceeded: {memory_usage:.2f} MB used; limit is {max_memory} MB.")
            os._exit(1)
        if elapsed_time > 30:
            # print(f"Run {elapsed_time: .0f}s, Current memory usage: {memory_usage:.2f} MB.")
            pass
        time.sleep(check_interval)

def start_memory_watcher(max_memory, check_interval=1):
    """
    Start the memory watcher in a daemon thread.
    :param max_memory: Maximum memory in MB.
    :param check_interval: Time interval in seconds between memory checks.
    """
    watcher_thread = threading.Thread(target=memory_watcher, args=(max_memory, check_interval), daemon=True)
    watcher_thread.start()

def auto_adjust_excel_dimensions(excel_path):
    """
    Adjust the column width and row height in an Excel file to fit the content.

    :param excel_path: Path to the Excel file to adjust.
    """
    wb = load_workbook(excel_path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column
            for cell in col:
                try:
                    lines = str(cell.value).split('\n')
                    longest_line = max(lines, key=len, default="")
                    if len(longest_line) > max_length:
                        max_length = len(longest_line)
                except:
                    pass
            adjusted_width = (max_length + 2)  # 添加一点额外空间
            ws.column_dimensions[get_column_letter(column)].width = adjusted_width

        # Adjust row heights
        for row in ws.iter_rows():
            max_line_count = 1
            for cell in row:
                try:
                    line_count = str(cell.value).count('\n') + 1
                    if line_count > max_line_count:
                        max_line_count = line_count
                except:
                    pass
            row_height = 15 * max_line_count # 假设标准单行高度为15
            ws.row_dimensions[row[0].row].height = row_height

    wb.save(excel_path)

def get_p4_workspace_root():
    # 获取当前Perforce客户端的信息
    p4_info_cmd = ["p4", "info"]
    p4_info_result = subprocess.run(p4_info_cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)

    if p4_info_result.returncode != 0:
        print(f"Error: {p4_info_result.stderr}")
        return None

    # 解析输出以找到根路径
    root_match = re.search(r'^Client root: (.+)$', p4_info_result.stdout, re.MULTILINE)
    if root_match:
        return root_match.group(1)
    else:
        print("Could not determine the Perforce workspace root.")
        return None

def get_p4_file_revision_info(workspace_path, file_name, line_number, ctx_len):
    # 构建p4 files命令来查找所有匹配的文件
    p4_files_cmd = ["p4", "files", f"{workspace_path}/.../{file_name}"]
    files_result = subprocess.run(p4_files_cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)

    if files_result.returncode != 0:
        print(f"Error: {files_result.stderr}")
        return None

    # 解析命令输出以找到所有匹配的文件
    matching_files = re.findall(r'([^#\s]+)#\d+ - ', files_result.stdout)

    # 对于每个匹配的文件，使用p4 annotate来找到指定行的最后修改人
    results = []
    for file_path in matching_files:
        annotate_cmd = ["p4", "annotate", "-u", "-c", "-I", "-t", file_path]
        annotate_result = subprocess.run(annotate_cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE)

        if annotate_result.returncode == 0:
            # 解析输出以找到指定行号的最后修改人和修订版本
            try:
                text_output = annotate_result.stdout.decode('utf-8')
            except UnicodeDecodeError:
                text_output = ""
                pass
            lines = text_output.splitlines()
            for idx, line in enumerate(lines):
                if idx == int(line_number):  # 第0行是文件路径
                    result = {
                        'path' : file_path,
                    }
                    # 上下文信息：-1不打印、0本行、n前后各n行
                    if ctx_len >= 0:
                        start, end = max(0, idx - ctx_len), min(len(lines), idx + ctx_len + 1)
                        result['context'] = lines[start:end]
                    # 分析具体的author等
                    match = re.match(r'^(\d+):\s+(\S+)\s+(\S+)\s+(.*)$', line)
                    if match:
                        change, author, date, content = match.groups()
                        result['change'] = change
                        result['author'] = author
                        result['date'] = date
                        result['content'] = content
                    results.append(result)
                    break
    return results

def extract_bracket_contents(log_line, left='[', right=']'):
    parts = []
    stack = []
    temp = ""
    for char in log_line:
        if char == left:
            if stack:
                temp += char
            stack.append(char)
        elif char == right:
            if len(stack) == 1:
                parts.append(temp)
                temp = ""
            else:
                temp += char
            stack.pop()
        else:
            if stack:
                temp += char
    # 如果栈不为空，说明有未闭合的括号
    if stack:
        raise ValueError("Unmatched brackets in the string")
    return parts

def parse_timestamp(log):
    # 2024-04-01 20:54:53.415 或 2024-04-02T20:46:23.415+0800
    for fmt in ('%Y-%m-%d %H:%M:%S.%f', '%Y-%m-%dT%H:%M:%S.%f%z'):
        try:
            dt = datetime.strptime(log, fmt)
            # 如果 datetime 对象是 naive 的，将其设置为 UTC 时区
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            return dt
        except ValueError:
            continue
    # 如果都不匹配，返回一个特定的默认时间戳
    return datetime(1970, 1, 1, tzinfo=timezone.utc)

def parse_significant_digits(log, significant_digits):
    # 四舍五入到指定有效位
    pattern = re.compile(r'([-+]?\d*\.\d+|[-+]?\d+)')
    match = pattern.search(log)
    if match:
        number_str = match.group(0)
        number_decimal = Decimal(number_str)
        # 移动小数点使得只有有效位数的数字在小数点前
        shift = number_decimal.adjusted() - significant_digits + 1
        # 四舍五入到有效位数
        rounded_number = number_decimal.scaleb(-shift).quantize(Decimal('1'), rounding=ROUND_HALF_UP).scaleb(shift)
        start, end = match.span(0)
        log = log[:start] + str(rounded_number) + log[end:]
    return log

def parse_decimal_places(log, decimal_places):
    # 四舍五入到指定小数位
    pattern = re.compile(r'([-+]?\d*\.\d+|[-+]?\d+)')
    match = pattern.search(log)
    if match:
        number_str = match.group(0)
        rounded_number = Decimal(number_str).quantize(Decimal('1.' + '0' * decimal_places), rounding=ROUND_HALF_UP)
        start, end = match.span(0)
        log = log[:start] + str(rounded_number) + log[end:]
    return log

def get_nested_json_value(data, path, default="null"):
    """安全地获取嵌套的 JSON 数据，包括数组中的值"""
    keys = path.replace('[', '.[').split('.')  # 将方括号前的点添加进来以便分割
    for key in keys:
        if key.startswith('[') and key.endswith(']') and isinstance(data, list):
            # 去掉方括号并尝试将索引转换为整数
            try:
                index = int(key[1:-1])
                data = data[index]
            except (ValueError, IndexError):
                # 索引无效或超出范围
                return default
        elif isinstance(data, dict) and key in data:
            data = data[key]
        else:
            return default
    return data

# 预处理函数
def preprocess_log(log_line, field_mapping_def):
    log_parts = {}
    if field_mapping_def['seperator'] == 'none':
        # WHOLE 定义，整个逻辑行是一个字段
        log_parts = {'MESSAGE': log_line.strip()}
    elif field_mapping_def['seperator'] == 'brackets':
        if not log_line.startswith('[') or not log_line.strip().endswith(']'):
            parts = []
        else:
            parts = extract_bracket_contents(log_line)
            # parts = re.findall(r'\[(?:[^\[\]]|\[[^\[\]]*\])*\]', log_line)
            # parts = re.findall(r'\[([^[\]]+(?:\[[^\]]*\])?[^[\]]*)\]', log_line)
        for key, index in field_mapping_def['mapping'].items():
            try:
                log_parts[key] = parts[index - 1]
            except IndexError:
                log_parts[key] = "null"
    elif field_mapping_def['seperator'] == 'json':
        try:
            log_json = json.loads(log_line)
        except json.JSONDecodeError:
            log_json = {}
        for key, field in field_mapping_def['mapping'].items():
            log_parts[key] = get_nested_json_value(log_json, field)
    else:
        parts = log_line.strip().split(field_mapping_def['seperator'])
        for key, index in field_mapping_def['mapping'].items():
            try:
                log_parts[key] = parts[index - 1]
            except IndexError:
                log_parts[key] = "null"
    return log_parts

# 聚类函数
def group_log(logs, merge_def):
    groups = []
    group_info = defaultdict(list)
    group_indexes = defaultdict(list)
    if not logs:
        return groups, group_info, group_indexes
    if merge_def == 'DBSCAN':
        vectorizer = TfidfVectorizer()
        X = vectorizer.fit_transform(logs)
        dbscan = DBSCAN(eps=0.5, min_samples=5)
        dbscan.fit(X)
        groups = dbscan.labels_
        # 为DBSCAN生成簇信息和索引
        for idx, label in enumerate(groups):
            group_indexes[label].append(idx)
            if label != -1:
                group_info[label].append(logs[idx])
        for label, log_list in group_info.items():
            if label != -1:
                group_info[label] = log_list[:1]  # top1
    elif merge_def.startswith('K_MEANS'):
        num_groups = int(merge_def.split('-')[1])
        vectorizer = TfidfVectorizer()
        X = vectorizer.fit_transform(logs)
        num_groups = min(num_groups, X.shape[0])
        kmeans = KMeans(n_clusters=num_groups, random_state=42)
        kmeans.fit(X)
        groups = kmeans.labels_
        # 为K_MEANS生成簇信息和索引
        for i in range(num_groups):
            group_indexes[i] = list(np.where(groups == i)[0])
            centroid_vector = kmeans.cluster_centers_[i]
            log_vectors = X[groups == i]
            if log_vectors.shape[0] > 0:
                similarities = cosine_similarity(centroid_vector.reshape(1, -1), log_vectors)
                top_index = similarities.argsort()[0][-1:]  # top1
                for index in top_index:
                    original_log_index = group_indexes[i][index]
                    group_info[i].append(logs[original_log_index])
    elif merge_def.startswith('TIME'):
        time_interval = int(merge_def.split('-')[1])
        timestamps = [parse_timestamp(log) for log in logs]
        if timestamps:
            min_time = min(timestamps)
            min_time_rounded = min_time.replace(minute=0, second=0, microsecond=0)
            groups = [int((timestamp - min_time_rounded).total_seconds() // time_interval) for timestamp in timestamps]
            # 为TIME生成簇信息和索引
            for idx, group_id in enumerate(groups):
                group_indexes[group_id].append(idx)
                start_time = min_time_rounded + timedelta(seconds=group_id * time_interval)
                end_time = start_time + timedelta(seconds=time_interval)
                group_info[group_id] = {
                    'start': start_time.strftime('%Y-%m-%d %H:%M:%S.%f'),
                    'end': end_time.strftime('%Y-%m-%d %H:%M:%S.%f')
                }
    elif merge_def.startswith('ROUND_DECIMAL'):
        decimal_places = int(merge_def.split('-')[1])
        values = [parse_decimal_places(log, decimal_places) for log in logs]
        if values:
            unique_logs = list(set(values))
            groups = [unique_logs.index(log) for log in values]
            # 为ROUND_DECIMAL生成簇信息和索引
            for idx, log in enumerate(values):
                group_id = unique_logs.index(log)
                group_indexes[group_id].append(idx)
            group_info = {i: log for i, log in enumerate(unique_logs)}
    elif merge_def.startswith('ROUND_SIGNIFICANT'):
        decimal_places = int(merge_def.split('-')[1])
        values = [parse_significant_digits(log, decimal_places) for log in logs]
        if values:
            unique_logs = list(set(values))
            groups = [unique_logs.index(log) for log in values]
            # 为ROUND_SIGNIFICANT生成簇信息和索引
            for idx, log in enumerate(values):
                group_id = unique_logs.index(log)
                group_indexes[group_id].append(idx)
            group_info = {i: log for i, log in enumerate(unique_logs)}
    elif merge_def == 'EQUAL':
        unique_logs = list(set(logs))
        groups = [unique_logs.index(log) for log in logs]
        # 为EQUAL生成簇信息和索引
        for idx, log in enumerate(logs):
            group_id = unique_logs.index(log)
            group_indexes[group_id].append(idx)
        group_info = {i: log for i, log in enumerate(unique_logs)}
    elif merge_def == 'NONE':
        # 对于 NONE 类型，不归类，每个日志条目都是自己的类别
        groups = list(range(len(logs)))
        for idx, log in enumerate(logs):
            group_indexes[idx].append(idx)
            group_info[idx] = [log]
    else:
        raise ValueError(f"Unsupported merge definition: {merge_def}")
    # 返回聚类结果和每个簇的信息和索引
    return groups, group_info, group_indexes

# 主处理函数
def process_log(input_config, output_config, define):
    if output_config['merge_file']:
        preprocessed_logs = []
        for file_key in input_config['file_filter'].keys():
            process_input(preprocessed_logs, input_config, define, file_key)
        process_output(preprocessed_logs, output_config, define, "MERGED")
    else:
        for file_key in input_config['file_filter'].keys():
            preprocessed_logs = []
            process_input(preprocessed_logs, input_config, define, file_key)
            process_output(preprocessed_logs, output_config, define, file_key)

def process_input(preprocessed_logs, input_config, define, file_key, use_multithreading=False):
    file_paths = []
    for path in input_config['file_filter'][file_key]:
        path = os.path.join(input_config['root_path'], path)
        matched_files = glob.glob(path)
        matched_files = [f for f in matched_files if not f.endswith('.gz')]
        matched_files = [f for f in matched_files if not f.endswith('.tgz')]
        if matched_files:
            file_paths.extend(matched_files)
            print(f"Include: {matched_files}")
        else:
            print(f"Warning: No files found for pattern {path}")

    field_mapping_def = define["field_mapping_def"][input_config['field_mapping']]

    # 内部函数，用于处理单个文件
    def process_single_file(log_file, line_pattern, line_filter, line_extract, field_mapping_def, field_filters):
        preprocessed_data = []
        with open(log_file, 'r', encoding='utf-8', errors='ignore') as file:
            file_content = file.read()
            # 使用正则表达式匹配逻辑行
            logical_lines = re.findall(line_pattern, file_content, re.MULTILINE | re.DOTALL)
            for log in logical_lines:
                # 如果 line_filter 不是 ".*"，则应用 line_filter 正则匹配
                if line_filter != ".*" and not re.search(line_filter, log):
                    continue
                # 如果 line_extract 不是 ".*"，则应用 line_extract 正则匹配
                if line_extract != ".*":
                    extracted_data = re.search(line_extract, log)
                    if extracted_data:
                        log = extracted_data.group(0)
                    else:
                        continue
                log_data = preprocess_log(log, field_mapping_def)
                # 应用 field_filters
                if all(re.search(f['match'], log_data.get(f['field'], '')) for f in field_filters):
                    preprocessed_data.append(log_data)
            print(f"file:{log_file}, logical_lines:{len(logical_lines)}, matched_lines:{len(preprocessed_data)}")
        return preprocessed_data

    line_pattern    = input_config.get('line_pattern', r'^(.*?)(?=\n|$)')   #默认行为类似:readline
    line_filter     = input_config.get('line_filter', r'.*')                #默认使用所有逻辑行
    line_extract    = input_config.get('line_extract', r'.*')               #默认使用该逻辑行的所有内容
    field_filters   = input_config.get('field_filters', [])                 #默认[], 不会根据field过滤

    if use_multithreading:
        # 使用多线程处理日志文件
        with ThreadPoolExecutor() as executor:
            future_to_file_data = {executor.submit(process_single_file, file_path, line_pattern, line_filter, line_extract, field_mapping_def, field_filters): file_path for file_path in file_paths}
            for future in as_completed(future_to_file_data):
                file_path = future_to_file_data[future]
                try:
                    data = future.result()
                    preprocessed_logs.extend(data)
                except Exception as exc:
                    print(f'File {file_path} generated an exception: {exc}')
    else:
        # 不使用多线程，直接顺序处理
        for file_path in file_paths:
            try:
                data = process_single_file(file_path, line_pattern, line_filter, line_extract, field_mapping_def, field_filters)
                preprocessed_logs.extend(data)
            except Exception as exc:
                print(f'File {file_path} generated an exception: {exc}')

def process_output(preprocessed_logs, output_config, define, file_key, use_multithreading=False):
    print(f"Processing {file_key}, total lines {len(preprocessed_logs)} ...")

    # 创建输出目录
    # 检查并创建 output/{args.conf} 目录
    output_dir = os.path.join('output', os.path.basename(args.conf))
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # 预处理结果，写入Excel文件
    if output_config.get('write_log_file', False):
        file_path = os.path.join(output_dir, f"{file_key}.txt")
        with open(file_path, 'w', encoding='utf-8', errors='ignore') as file:
            for log in preprocessed_logs:
                file.write(f"{log}\n")
        excel_file_path = os.path.join(output_dir, f"{file_key}.xlsx")
        excel_data = []
        for log in preprocessed_logs:
            excel_data.append(log)
        df = pd.DataFrame(excel_data)
        df.to_excel(excel_file_path, index=False)
        auto_adjust_excel_dimensions(excel_file_path)
        print(f"Preprocess Results have been written to {excel_file_path}")

    # 准备数据列表，用于最后写入Excel
    excel_data = []

    def process_toplist_item(toplist_item):
        field = toplist_item['field']
        count = toplist_item['count']
        merge = toplist_item['merge']

        field_logs = [log[field] for log in preprocessed_logs]
        return group_log(field_logs, merge), (file_key, count, merge, field)

    def print_results(results, file_key, count, merge, field):
        groups, group_info, group_indexes = results
        group_counter = Counter(groups)
        total_count = sum(group_counter.values())
        top_groups = group_counter.most_common(count)
        p4_root = get_p4_workspace_root()

        print(f'{file_key} Top {count} by {merge} for field {field}:')
        for idx, (group, group_count) in enumerate(top_groups):
            top_index = idx + 1
            percentage = (group_count / total_count) * 100
            content = group_info.get(group, "N/A")
            author, change, date, path, context = 'N/A', 'N/A', 'N/A', 'N/A', 'N/A'

            print(f'Count: {group_count}\t| Percentage: {percentage:.2f}%\t| Content: {content}'.encode('utf-8').decode('utf-8'))

            # 如果是CALLER，提取文件路径和行号，并打印对应的修改人
            if field == "CALLER" and output_config.get('p4_caller_info', False):
                caller_match = re.search(r'(\S+):(\d+)$', content)
                if caller_match:
                    file_path, line_number = caller_match.groups()
                    file_name = os.path.basename(file_path)
                    ctx_len = output_config.get('p4_caller_ctx_len', -1)
                    p4_info_list = get_p4_file_revision_info(p4_root, file_name, line_number, ctx_len)
                    if p4_info_list:
                        max_path_match_digits = 0
                        best_match_info = None
                        for p4_info in p4_info_list:
                            path = p4_info.get('path', 'null')
                            match_digits = sum(1 for i, j in zip(file_path[::-1], path[::-1]) if i == j)
                            # 更新最大匹配字符数和最佳匹配信息
                            if match_digits > max_path_match_digits:
                                max_path_match_digits = match_digits
                                best_match_info = p4_info
                        if best_match_info:
                            author = best_match_info.get('author', 'null')
                            change = best_match_info.get('change', 'null')
                            date = best_match_info.get('date', 'null')
                            context = "\n".join(best_match_info.get('context', []))
                            path = best_match_info.get('path', 'null')
                        print(f"Author: {author}\t| Date: {date}\t| Change: {change}\t| File Path: {path} \n{context}")
                    else:
                        print('Author information not available')

            # 根据配置获取示例日志
            print_line_num = output_config.get('print_line_num', 0)
            example_indexes = group_indexes[group][:print_line_num]
            examples = []
            for i in example_indexes:
                if 'print_line_field' in output_config:
                    example = {field: preprocessed_logs[i].get(field, 'N/A') for field in output_config['print_line_field']}
                    example_str = ", ".join(f"{key}: {value}" for key, value in example.items())
                    examples.append(example_str)
                else:
                    log_entry_str = ", ".join(f"{key}: {value}" for key, value in preprocessed_logs[i].items())
                    examples.append(log_entry_str)
            for log_str in examples:
                print("\t", log_str)
            
            # 添加数据到Excel数据列表
            excel_data.append({
                'FileKey': file_key,
                'Field': field,
                'Merge': merge,
                'TopIndex': top_index,
                'TopTotal': count,
                'Count': group_count,
                'Percentage': percentage,
                'Content': content,
                'Date': date,
                'Author': author,
                'Change': change,
                'Context': context,
                'Examples': "\n".join(examples)
            })

        print('-' * 80)

    if use_multithreading:
        with ProcessPoolExecutor() as executor:
            futures = [executor.submit(process_toplist_item, toplist_item) for toplist_item in output_config['toplist']]
            for future in as_completed(futures):
                results, info = future.result()
                print_results(results, *info)
    else:
        for toplist_item in output_config['toplist']:
            results, info = process_toplist_item(toplist_item)
            print_results(results, *info)

    # 数据统计结果，写入Excel文件
    df = pd.DataFrame(excel_data)
    excel_file_path = os.path.join(output_dir, f"{file_key}.result.xlsx")
    df.to_excel(excel_file_path, index=False)
    auto_adjust_excel_dimensions(excel_file_path)
    print(f"Statistic Results have been written to {excel_file_path}")

    # 打开输出目录
    if platform.system().lower() in ("win", "windows") and os.path.exists(output_dir):
        os.system("start " + output_dir)


if __name__ == "__main__":
    # 设置命令行参数解析
    parser = argparse.ArgumentParser(description='Log Statistics Processor')
    parser.add_argument('--conf', default='config/_example.json', help='Path to the configuration file')
    parser.add_argument('--mem_limit', default=10240, type=int, help='Memory limit in megabytes')
    args = parser.parse_args()

    # 设置内存限制(MB)
    if args.mem_limit:
        start_memory_watcher(args.mem_limit)

    # 读取配置文件路径
    config_file_path = args.conf

    # 读取配置文件
    try:
        with open(config_file_path, 'r', encoding='utf-8') as f:
            config = json.load(f)
    except FileNotFoundError:
        print(f"Error: Configuration file '{config_file_path}' not found.")
        exit(1)

    # 读取定义文件
    try:
        with open('define.json', 'r', encoding='utf-8') as f:
            define = json.load(f)
    except FileNotFoundError:
        print("Error: Definition file 'define.json' not found.")
        exit(1)

    # 运行主处理函数
    process_log(config['input'], config['output'], define)